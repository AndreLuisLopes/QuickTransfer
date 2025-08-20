import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook

COR_FUNDO = "#F2F2F2"
COR_FRAME = "#FFFFFF"
COR_LABEL = "#003366"
COR_BOTAO = "#0099CC"
COR_TEXTO = "#003366"

caminho_origem = ""
caminho_destino = ""
df_origem = pd.DataFrame()
abas_disponiveis = []

def selecionar_arquivo_origem():
    global caminho_origem, df_origem
    caminho_origem = filedialog.askopenfilename(
        title="Selecione o Arquivo de Origem",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )
    if not caminho_origem:
        return

    try:
        df_origem = pd.read_excel(caminho_origem)
        colunas = df_origem.columns.tolist()
        caixa_origem.delete("1.0", tk.END)
        caixa_origem.insert(tk.END, "\n".join(colunas))
        combobox_origem["values"] = colunas
        if colunas:
            combobox_origem.set(colunas[0])
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo de origem:\n{e}")

def selecionar_arquivo_destino():
    global caminho_destino, abas_disponiveis
    caminho_destino = filedialog.askopenfilename(
        title="Selecione o Arquivo de Destino",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )
    if not caminho_destino:
        return

    try:
        wb_temp = load_workbook(caminho_destino, read_only=True)
        abas_disponiveis = wb_temp.sheetnames
        wb_temp.close()

        combobox_abas["values"] = abas_disponiveis
        if abas_disponiveis:
            combobox_abas.set(abas_disponiveis[0])
            atualizar_colunas_destino()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo de destino:\n{e}")

def atualizar_colunas_destino(event=None):
    aba = combobox_abas.get()
    if not caminho_destino or not aba:
        return

    try:
        wb = load_workbook(caminho_destino, read_only=True)
        ws = wb[aba]
        cabecalhos = [
            str(cell.value).strip()
            for cell in ws[1]
            if cell.value
        ]
        wb.close()

        combobox_destino1["values"] = cabecalhos
        combobox_destino2["values"] = cabecalhos
        if cabecalhos:
            combobox_destino1.set(cabecalhos[0])
            combobox_destino2.set(cabecalhos[0])

        caixa_destino.delete("1.0", tk.END)
        caixa_destino.insert(tk.END, "\n".join(cabecalhos))
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao obter colunas da aba '{aba}':\n{e}")

def atualizar_interface_separacao():
    if var_separar.get():
        label_separador.grid(row=9, column=0, columnspan=2, pady=2, sticky="n")
        entry_separador.grid(row=10, column=0, columnspan=2, pady=2, sticky="n")
        label_destino2.grid(row=11, column=0, columnspan=2, pady=2, sticky="n")
        combobox_destino2.grid(row=12, column=0, columnspan=2, pady=2, sticky="n")
        botao_transferir.grid(row=13, column=0, columnspan=2, pady=15, sticky="n")
    else:
        label_separador.grid_forget()
        entry_separador.grid_forget()
        label_destino2.grid_forget()
        combobox_destino2.grid_forget()
        botao_transferir.grid(row=8, column=0, columnspan=2, pady=15, sticky="n")

def transferir_dados():
    origem_coluna = combobox_origem.get()
    filtro = entry_filtro.get().strip()
    separar = var_separar.get()
    separador = entry_separador.get().strip()
    destino1 = combobox_destino1.get()
    destino2 = combobox_destino2.get() if separar else None

    if not origem_coluna or not destino1:
        messagebox.showwarning(
            "Atenção",
            "Selecione a coluna de origem e pelo menos uma coluna de destino."
        )
        return

    try:
        dados = df_origem[origem_coluna].astype(str)
        if filtro:
            dados = dados[dados.str.contains(filtro, na=False)]
        dados = dados.reset_index(drop=True)
        if dados.empty:
            messagebox.showwarning(
                "Nenhum dado encontrado",
                "O filtro aplicado não retornou nenhum valor."
            )
            return

        mostrar_previa(dados, destino1, destino2, separador)
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao preparar os dados:\n{e}")

def mostrar_previa(dados, destino1, destino2, separador):
    previa = tk.Toplevel(janela)
    previa.title("Prévia dos Dados")
    previa.geometry("600x400")
    previa.configure(bg=COR_FUNDO)

    texto = tk.Text(previa, wrap="none", font=("Segoe UI", 10), bg=COR_FRAME, fg=COR_LABEL)
    texto.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

    texto.insert(tk.END, f"Coluna de origem: {combobox_origem.get()}\n")
    texto.insert(
        tk.END,
        f"Filtro aplicado: {entry_filtro.get().strip() or 'Nenhum'}\n"
    )
    texto.insert(tk.END, f"Separador: {separador or 'Nenhum'}\n\n")

    for i, valor in enumerate(dados, start=1):
        if destino2 and separador and separador in valor:
            p1, p2 = valor.split(separador, 1)
            linha = (
                f"{i:03}: [{destino1}] {p1.strip()} | "
                f"[{destino2}] {p2.strip()}\n"
            )
        else:
            linha = f"{i:03}: [{destino1}] {valor.strip()}\n"
        texto.insert(tk.END, linha)

    frame_botoes = ttk.Frame(previa, style="TFrame")
    frame_botoes.grid(row=1, column=0, pady=10, padx=10)

    def confirmar():
        previa.destroy()
        cont = executar_transferencia(dados, destino1, destino2, separador)
        messagebox.showinfo(
            "Transferência completa",
            f"{cont} células transferidas com sucesso!"
        )

    def cancelar():
        previa.destroy()

    ttk.Button(
        frame_botoes,
        text="Confirmar e Transferir",
        command=confirmar,
        style="TButton"
    ).grid(row=0, column=0, padx=10, pady=5)

    ttk.Button(
        frame_botoes,
        text="Cancelar",
        command=cancelar,
        style="TButton"
    ).grid(row=0, column=1, padx=10, pady=5)

    previa.grid_rowconfigure(0, weight=1)
    previa.grid_columnconfigure(0, weight=1)

def executar_transferencia(dados, dest1, dest2, sep):
    aba = combobox_abas.get()
    cont = 0
    try:
        wb = load_workbook(caminho_destino)
        ws = wb[aba]

        cab = [
            str(cell.value).strip() if cell.value else ""
            for cell in ws[1]
        ]
        idx1 = cab.index(dest1) + 1
        idx2 = cab.index(dest2) + 1 if dest2 else None

        for row, val in enumerate(dados, start=2):
            if dest2 and sep and sep in val:
                a, b = val.split(sep, 1)
                ws.cell(row=row, column=idx1, value=a.strip())
                ws.cell(row=row, column=idx2, value=b.strip())
                cont += 2
            else:
                ws.cell(row=row, column=idx1, value=val.strip())
                cont += 1

        wb.save(caminho_destino)
    except Exception as e:
        messagebox.showerror("Erro", f"Falha na transferência:\n{e}")
    return cont

janela = tk.Tk()
janela.title("QuickTransfer")
janela.geometry("900x700")
janela.configure(bg=COR_FUNDO)

style = ttk.Style(janela)
style.theme_use("clam")
style.configure("TFrame", background=COR_FRAME)
style.configure("TLabel", background=COR_FRAME, foreground=COR_LABEL, font=("Segoe UI", 11))
style.configure("TButton", background=COR_BOTAO, foreground=COR_FRAME, font=("Segoe UI", 11, "bold"))
style.configure("TCheckbutton", background=COR_FRAME, foreground=COR_LABEL, font=("Segoe UI", 11))
style.map("TButton", background=[("active", COR_LABEL)])

frame_principal = ttk.Frame(janela, style="TFrame")
frame_principal.grid(row=0, column=0, padx=30, pady=30, sticky="nsew")
janela.grid_rowconfigure(0, weight=1)
janela.grid_columnconfigure(0, weight=1)

frame_principal.grid_columnconfigure(0, weight=1)
frame_principal.grid_columnconfigure(1, weight=1)
frame_principal.grid_rowconfigure(0, weight=1)

frame_origem = ttk.LabelFrame(frame_principal, text="Arquivo de Origem")
frame_origem.grid(row=0, column=0, padx=15, pady=10, sticky="n")
ttk.Button(
    frame_origem,
    text="Selecionar Origem",
    command=selecionar_arquivo_origem,
    style="TButton"
).grid(row=0, column=0, pady=5)
caixa_origem = tk.Text(
    frame_origem, height=10, width=40, font=("Consolas", 10), bg=COR_FUNDO, fg=COR_LABEL
)
caixa_origem.grid(row=1, column=0, pady=5)
combobox_origem = ttk.Combobox(frame_origem, state="readonly")
combobox_origem.grid(row=2, column=0, pady=5)

frame_destino = ttk.LabelFrame(frame_principal, text="Arquivo de Destino")
frame_destino.grid(row=0, column=1, padx=15, pady=10, sticky="n")
ttk.Button(
    frame_destino,
    text="Selecionar Destino",
    command=selecionar_arquivo_destino,
    style="TButton"
).grid(row=0, column=0, pady=5)
caixa_destino = tk.Text(
    frame_destino, height=10, width=40, font=("Consolas", 10), bg=COR_FUNDO, fg=COR_LABEL
)
caixa_destino.grid(row=1, column=0, pady=5)

ttk.Label(frame_principal, text="Filtro (opcional):", style="TLabel").grid(row=1, column=0, columnspan=2, pady=2, sticky="n")
entry_filtro = ttk.Entry(frame_principal, width=30)
entry_filtro.grid(row=2, column=0, columnspan=2, pady=2, sticky="n")

ttk.Label(frame_principal, text="Selecione a aba da planilha de destino:", style="TLabel").grid(row=3, column=0, columnspan=2, pady=2, sticky="n")
combobox_abas = ttk.Combobox(frame_principal, state="readonly")
combobox_abas.grid(row=4, column=0, columnspan=2, pady=2, sticky="n")
combobox_abas.bind("<<ComboboxSelected>>", atualizar_colunas_destino)

var_separar = tk.BooleanVar(value=False)
ttk.Checkbutton(
    frame_principal,
    text="Separar valores",
    variable=var_separar,
    command=atualizar_interface_separacao,
    style="TCheckbutton"
).grid(row=5, column=0, columnspan=2, pady=2, sticky="n")

label_separador = ttk.Label(frame_principal, text="Caracter separador:", style="TLabel")
entry_separador = ttk.Entry(frame_principal)

ttk.Label(frame_principal, text="Destino da primeira parte:", style="TLabel").grid(row=6, column=0, columnspan=2, pady=2, sticky="n")
combobox_destino1 = ttk.Combobox(frame_principal, state="readonly")
combobox_destino1.grid(row=7, column=0, columnspan=2, pady=2, sticky="n")

label_destino2 = ttk.Label(frame_principal, text="Destino da segunda parte:", style="TLabel")
combobox_destino2 = ttk.Combobox(frame_principal, state="readonly")

botao_transferir = ttk.Button(
    frame_principal,
    text="Transferir Dados",
    command=transferir_dados,
    style="TButton"
)
botao_transferir.grid(row=8, column=0, columnspan=2, pady=15, sticky="n")

atualizar_interface_separacao()

janela.mainloop()